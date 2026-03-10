#!/usr/bin/env python3
"""
WGS Fold Penalty Calculator - Multi-file Version

This version can process multiple Picard files at once and automatically
matches them to samples in the mosdepth file.

Usage:
    python calculate_wgs_fold_penalties_multi.py \
        -m mosdepth.tsv \
        -p file1.wgs.txt file2.wgs.txt file3.wgs.txt \
        --excel results.xlsx
"""

import pandas as pd
import numpy as np
import argparse
import sys
import re
from pathlib import Path


def extract_sample_from_filename(filename):
    """
    Extract sample name from filename.
    
    Examples:
        24300FL-13-01-01_S1_L005.wgs.txt -> 24300FL-13-01-01_S1_L005
        sample1.wgs.txt -> sample1
    """
    basename = Path(filename).stem
    # Remove common suffixes
    for suffix in ['.wgs', '.metrics', '_wgs_metrics']:
        basename = basename.replace(suffix, '')
    return basename


def read_picard_wgs_metrics_txt(picard_file):
    """
    Read Picard CollectWgsMetrics text output.
    Returns dict with sample info extracted from filename.
    """
    # Try to extract sample name from filename
    sample_name = extract_sample_from_filename(picard_file)
    
    with open(picard_file, 'r') as f:
        lines = f.readlines()
    
    # Find the metrics header
    header_idx = None
    for i, line in enumerate(lines):
        if 'GENOME_TERRITORY' in line or 'MEAN_COVERAGE' in line:
            header_idx = i
            break
    
    if header_idx is None:
        raise ValueError(f"Could not find metrics header in {picard_file}")
    
    # Parse header
    headers = lines[header_idx].strip().split('\t')
    
    # Find column indices
    try:
        mean_idx = headers.index('MEAN_COVERAGE')
        median_idx = headers.index('MEDIAN_COVERAGE')
    except ValueError:
        raise ValueError(f"Could not find MEAN_COVERAGE or MEDIAN_COVERAGE in {picard_file}")
    
    # Parse data line (first non-empty line after header)
    for line in lines[header_idx + 1:]:
        if line.strip():
            parts = line.strip().split('\t')
            if len(parts) > max(mean_idx, median_idx):
                mean_cov = float(parts[mean_idx])
                median_cov = float(parts[median_idx])
                return {
                    'sample': sample_name,
                    'mean': mean_cov,
                    'median': median_cov,
                    'source_file': picard_file
                }
    
    raise ValueError(f"Could not parse data from {picard_file}")


def read_multiple_picard_files(picard_files):
    """
    Read multiple Picard files and return combined dict.
    """
    picard_data = {}
    
    for pfile in picard_files:
        try:
            data = read_picard_wgs_metrics_txt(pfile)
            sample = data['sample']
            picard_data[sample] = {
                'mean': data['mean'],
                'median': data['median'],
                'source_file': data['source_file']
            }
            print(f"  ✓ Loaded {sample} from {Path(pfile).name}")
        except Exception as e:
            print(f"  ✗ Error reading {pfile}: {e}")
    
    return picard_data


def read_mosdepth_cumulative(mosdepth_file):
    """Read mosdepth cumulative coverage distribution file."""
    df = pd.read_csv(mosdepth_file, sep='\t')
    df.columns = [col.replace('\ufeff', '') for col in df.columns]
    return df


def calculate_percentile_coverage(coverages, percentages, target_pct):
    """Calculate exact coverage at a given percentile using linear interpolation."""
    for i in range(len(percentages) - 1):
        if percentages[i] >= target_pct and percentages[i+1] < target_pct:
            cov1, pct1 = coverages[i], percentages[i]
            cov2, pct2 = coverages[i+1], percentages[i+1]
            coverage = cov1 + (target_pct - pct1) / (pct2 - pct1) * (cov2 - cov1)
            return coverage
        elif percentages[i] == target_pct:
            return coverages[i]
    return None


def calculate_fold_penalties(mosdepth_df, picard_data):
    """Calculate Fold-80 and Fold-90 penalties for all samples."""
    results = []
    
    for idx, row in mosdepth_df.iterrows():
        if pd.isna(row['Sample']):
            continue
        
        sample = row['Sample']
        
        # Try to find matching sample in Picard data
        matched_sample = None
        if sample in picard_data:
            matched_sample = sample
        else:
            # Try partial matching
            for picard_sample in picard_data.keys():
                if picard_sample in sample or sample in picard_sample:
                    matched_sample = picard_sample
                    print(f"  Matched '{sample}' -> '{picard_sample}'")
                    break
        
        if not matched_sample:
            print(f"  Warning: {sample} not found in Picard data, skipping")
            continue
        
        # Extract coverage distribution
        coverages = []
        percentages = []
        
        for col in mosdepth_df.columns[1:]:
            try:
                coverage_depth = int(col)
                pct = row[col]
                if pd.notna(pct):
                    coverages.append(coverage_depth)
                    percentages.append(float(pct))
            except ValueError:
                continue
        
        coverages = np.array(coverages)
        percentages = np.array(percentages)
        
        # Get mean and median from Picard
        mean_cov = picard_data[matched_sample]['mean']
        median_cov = picard_data[matched_sample]['median']
        
        # Calculate Fold-80 and Fold-90
        cov_20th = calculate_percentile_coverage(coverages, percentages, 80.0)
        fold_80 = mean_cov / cov_20th if cov_20th else None
        
        cov_10th = calculate_percentile_coverage(coverages, percentages, 90.0)
        fold_90 = mean_cov / cov_10th if cov_10th else None
        
        results.append({
            'sample': sample,
            'mean_coverage': mean_cov,
            'median_coverage': median_cov,
            'coverage_20th_percentile': cov_20th,
            'fold_80_penalty': fold_80,
            'coverage_10th_percentile': cov_10th,
            'fold_90_penalty': fold_90,
            'picard_source': Path(picard_data[matched_sample]['source_file']).name
        })
    
    return results


def assess_quality(fold_80, fold_90):
    """Assess coverage quality based on fold penalties."""
    if fold_80 < 1.2:
        quality_80 = "Excellent"
    elif fold_80 < 1.5:
        quality_80 = "Good"
    elif fold_80 < 2.0:
        quality_80 = "Acceptable"
    else:
        quality_80 = "Poor"
    
    if fold_90 < 1.5:
        quality_90 = "Excellent"
    elif fold_90 < 2.0:
        quality_90 = "Good"
    elif fold_90 < 2.5:
        quality_90 = "Acceptable"
    else:
        quality_90 = "Poor"
    
    return quality_80, quality_90


def print_results(results):
    """Print results in a formatted table."""
    print("\n" + "="*110)
    print("WGS FOLD PENALTY ANALYSIS RESULTS")
    print("="*110)
    print("\nFormula: MEAN_COVERAGE / Coverage_at_Xth_percentile")
    print("Interpretation: How much additional sequencing needed to bring bottom X% to mean coverage")
    print("="*110)
    print()
    
    print(f"{'Sample':<35} {'Mean':<10} {'Fold-80':<12} {'Quality':<12} {'Fold-90':<12} {'Quality'}")
    print("-"*110)
    
    for r in results:
        quality_80, quality_90 = assess_quality(r['fold_80_penalty'], r['fold_90_penalty'])
        
        print(f"{r['sample']:<35} "
              f"{r['mean_coverage']:<10.2f} "
              f"{r['fold_80_penalty']:<12.4f} "
              f"{quality_80:<12} "
              f"{r['fold_90_penalty']:<12.4f} "
              f"{quality_90}")
    
    print("\n" + "="*110)
    print("QUALITY THRESHOLDS:")
    print("-"*110)
    print("Fold-80: <1.2 = Excellent | 1.2-1.5 = Good | 1.5-2.0 = Acceptable | >2.0 = Poor")
    print("Fold-90: <1.5 = Excellent | 1.5-2.0 = Good | 2.0-2.5 = Acceptable | >2.5 = Poor")
    print("="*110)
    print()


def save_results_csv(results, output_file):
    """Save results to CSV file."""
    df = pd.DataFrame(results)
    qualities = [assess_quality(r['fold_80_penalty'], r['fold_90_penalty']) for r in results]
    df['fold_80_quality'] = [q[0] for q in qualities]
    df['fold_90_quality'] = [q[1] for q in qualities]
    
    column_order = [
        'sample', 'mean_coverage', 'median_coverage',
        'coverage_20th_percentile', 'fold_80_penalty', 'fold_80_quality',
        'coverage_10th_percentile', 'fold_90_penalty', 'fold_90_quality',
        'picard_source'
    ]
    df = df[column_order]
    df.to_csv(output_file, index=False, float_format='%.4f')
    print(f"✓ Results saved to: {output_file}")


def save_results_excel(results, output_file):
    """Save results to Excel file with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Warning: openpyxl not installed. Cannot create Excel file.")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fold Penalties"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = "WGS Fold-80 and Fold-90 Penalty Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')
    
    ws['A2'] = "Formula: MEAN_COVERAGE ÷ Coverage_at_Xth_percentile"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:I2')
    
    headers = ["Sample", "Mean\nCoverage", "Median\nCoverage",
               "20th %ile\nCoverage", "Fold-80\nPenalty", "Quality",
               "10th %ile\nCoverage", "Fold-90\nPenalty", "Quality"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    row_num = 5
    for r in results:
        quality_80, quality_90 = assess_quality(r['fold_80_penalty'], r['fold_90_penalty'])
        
        ws.cell(row=row_num, column=1, value=r['sample']).border = border
        ws.cell(row=row_num, column=2, value=f"{r['mean_coverage']:.2f}").border = border
        ws.cell(row=row_num, column=3, value=f"{r['median_coverage']:.2f}").border = border
        ws.cell(row=row_num, column=4, value=f"{r['coverage_20th_percentile']:.2f}").border = border
        ws.cell(row=row_num, column=5, value=round(r['fold_80_penalty'], 4)).border = border
        
        fill_80 = PatternFill(start_color="C6EFCE" if quality_80 == "Excellent" else 
                             "FFEB9C" if quality_80 == "Good" else "FFC7CE", 
                             end_color="C6EFCE" if quality_80 == "Excellent" else 
                             "FFEB9C" if quality_80 == "Good" else "FFC7CE", 
                             fill_type="solid")
        cell = ws.cell(row=row_num, column=6, value=quality_80)
        cell.border, cell.fill = border, fill_80
        cell.alignment = Alignment(horizontal='center')
        
        ws.cell(row=row_num, column=7, value=f"{r['coverage_10th_percentile']:.2f}").border = border
        ws.cell(row=row_num, column=8, value=round(r['fold_90_penalty'], 4)).border = border
        
        fill_90 = PatternFill(start_color="C6EFCE" if quality_90 == "Excellent" else 
                             "FFEB9C" if quality_90 == "Good" else "FFC7CE",
                             end_color="C6EFCE" if quality_90 == "Excellent" else 
                             "FFEB9C" if quality_90 == "Good" else "FFC7CE",
                             fill_type="solid")
        cell = ws.cell(row=row_num, column=9, value=quality_90)
        cell.border, cell.fill = border, fill_90
        cell.alignment = Alignment(horizontal='center')
        
        row_num += 1
    
    row_num += 2
    ws.cell(row=row_num, column=1, value="Quality Thresholds:").font = Font(bold=True, size=11)
    row_num += 1
    ws.cell(row=row_num, column=1, value="Fold-80:").font = Font(bold=True)
    ws.cell(row=row_num, column=2, value="< 1.2 = Excellent")
    ws.cell(row=row_num, column=4, value="1.2-1.5 = Good")
    row_num += 1
    ws.cell(row=row_num, column=1, value="Fold-90:").font = Font(bold=True)
    ws.cell(row=row_num, column=2, value="< 1.5 = Excellent")
    ws.cell(row=row_num, column=4, value="1.5-2.0 = Good")
    
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C']:
        ws.column_dimensions[col].width = 12
    for col in ['D', 'G']:
        ws.column_dimensions[col].width = 14
    for col in ['E', 'H']:
        ws.column_dimensions[col].width = 12
    for col in ['F', 'I']:
        ws.column_dimensions[col].width = 16
    
    wb.save(output_file)
    print(f"✓ Excel results saved to: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Calculate WGS Fold penalties - handles multiple Picard files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single Picard file
  %(prog)s -m mosdepth.tsv -p sample1.wgs.txt --excel results.xlsx
  
  # Multiple Picard files (one per sample)
  %(prog)s -m mosdepth.tsv -p *.wgs.txt --excel results.xlsx
  
  # Specific files
  %(prog)s -m mosdepth.tsv -p s1.wgs.txt s2.wgs.txt s3.wgs.txt -o out.csv
        """
    )
    
    parser.add_argument('-m', '--mosdepth', required=True,
                       help='Mosdepth cumulative coverage file (.tsv)')
    parser.add_argument('-p', '--picard', nargs='+', required=True,
                       help='Picard CollectWgsMetrics file(s) (.txt)')
    parser.add_argument('-o', '--output',
                       help='Output CSV file (optional)')
    parser.add_argument('--excel',
                       help='Output Excel file (optional)')
    
    args = parser.parse_args()
    
    # Check mosdepth file
    if not Path(args.mosdepth).exists():
        print(f"Error: Mosdepth file not found: {args.mosdepth}")
        sys.exit(1)
    
    # Check Picard files
    valid_picard_files = []
    for pfile in args.picard:
        if Path(pfile).exists():
            valid_picard_files.append(pfile)
        else:
            print(f"Warning: Picard file not found: {pfile}")
    
    if not valid_picard_files:
        print("Error: No valid Picard files found")
        sys.exit(1)
    
    # Read files
    print(f"Reading mosdepth file: {args.mosdepth}")
    mosdepth_df = read_mosdepth_cumulative(args.mosdepth)
    print(f"  Found {len(mosdepth_df)} samples in mosdepth file")
    
    print(f"\nReading {len(valid_picard_files)} Picard file(s):")
    picard_data = read_multiple_picard_files(valid_picard_files)
    
    if not picard_data:
        print("\nError: Could not read any Picard files")
        sys.exit(1)
    
    print(f"\nCalculating fold penalties...")
    results = calculate_fold_penalties(mosdepth_df, picard_data)
    
    if not results:
        print("\nError: No results calculated.")
        print("\nTroubleshooting:")
        print("  - Check that sample names match between mosdepth and Picard files")
        print("  - Mosdepth samples:", list(mosdepth_df['Sample'].dropna()))
        print("  - Picard samples:", list(picard_data.keys()))
        sys.exit(1)
    
    print_results(results)
    
    if args.output:
        save_results_csv(results, args.output)
    
    if args.excel:
        save_results_excel(results, args.excel)
    
    print("\n✓ Analysis complete!")


if __name__ == '__main__':
    main()
